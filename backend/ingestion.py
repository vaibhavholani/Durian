"""
Ingestion module for Durian's Main Consolidated file.xlsx.

File structure:
  Row 1  – ownership annotations (Prachi, Check…)
  Row 2  – column numbering
  Rows 3-5 – multi-platform mappings (Amazon, Flipkart…)
  Row 6  – top-level category groups (Room, L3…)
  Row 7  – ACTUAL HEADER ROW used by Durian internally
  Row 8+ – product data

Also reads MiracleAi.xlsx for gold examples (approved hand-written content).
"""

import re
import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional, Tuple

from models import ProductInput


# ── Fixed column indices (from row 7 of the Consolidated file) ───────────────
# Verified by reading the actual file.
COL = {
    "room":                     0,    # "Room"
    "product_type":             2,    # "L3 / Type"
    "product_name":             5,    # "Model"
    "sku_id":                   6,    # "SKU"
    "imported":                 8,    # "Domestic/Imported"
    "small_description":        12,   # "Small Description"
    "design_story":             14,   # "Design Story"
    "meta_keywords":            16,   # "Meta Keyword"
    "bullet_points":            18,   # "Features / Bullet Point"
    "primary_color":            22,   # "Primary Colour"
    "secondary_color":          23,   # "Colour & Colour Family"
    "finish":                   31,   # "Finish"
    "primary_material":         32,   # "Primary Material"
    "filling_material":         37,   # "Filling Material"
    "upholstery_type":          38,   # "Upholstery Material"
    "leg_material":             46,   # "Leg Material"
    "configuration":            55,   # "Orientations"
    "dimensions":               63,   # "Size"
    "net_weight":               64,   # "Net Weight"
    "width_mm":                 75,   # "Width (mm)"
    "depth_mm":                 76,   # "Depth (mm)"
    "height_mm":                77,   # "Height (mm)"
    "max_load_capacity":        92,   # "Maximum Load Capacity (kg)"
    "mechanism":                96,   # "Mechanism"
    "function":                 97,   # "Function"
    "warranty":                 106,  # "Warranty"
    "style":                    108,  # "Style"
    "care_instructions":        114,  # "Care Instructions"
    "w_icon_1":                 127,  # "W Icon 1"
    "w_icon_2":                 128,  # "W Icon 2"
    "w_icon_3":                 129,  # "W Icon 3"
    "w_icon_4":                 130,  # "W Icon 4"
}

# Header row index (1-based = 7, pandas 0-based = 6)
HEADER_ROW_IDX = 6   # row 7 in Excel (0-indexed)
DATA_START_ROW  = 7  # row 8 in Excel (0-indexed)


def _v(row: tuple, col_idx: int) -> Optional[str]:
    """Safe value extractor: returns stripped string or None."""
    if col_idx >= len(row):
        return None
    val = row[col_idx]
    if val is None or (isinstance(val, float) and str(val) == 'nan'):
        return None
    s = str(val).strip()
    return s if s and s.lower() != 'nan' else None


def _parse_imported(val: Optional[str]) -> Optional[bool]:
    if not val:
        return None
    s = val.lower()
    if any(k in s for k in ["import", "yes", "true", "1"]):
        return True
    if any(k in s for k in ["domestic", "india", "local", "no", "false", "0"]):
        return False
    return None


def _find_header_row(ws) -> int:
    """
    Scan rows looking for a row that contains 'SKU' — that is the header row.
    Returns 0-based row index.
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        vals = [str(v).strip().lower() for v in row if v is not None]
        if "sku" in vals:
            return idx
    return HEADER_ROW_IDX  # fallback


def parse_consolidated_excel(file_path: str) -> List[ProductInput]:
    """
    Parse the Main Consolidated file (or any similarly structured file) into
    a list of ProductInput objects. Expects header at row 7 (0-based: 6).
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    # Try Sheet1 first, else first sheet
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]

    header_idx = _find_header_row(ws)
    data_start  = header_idx + 1

    products: List[ProductInput] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start + 1, values_only=True)):
        sku = _v(row, COL["sku_id"])
        name = _v(row, COL["product_name"])

        # Skip empty rows or rows that look like sub-headers
        if not sku or sku.lower() in ("sku", "model", "item name", ""):
            continue
        if name and name.lower() in ("model", "product name"):
            continue

        # Build dimensions string from individual mm columns if needed
        w = _v(row, COL["width_mm"])
        d = _v(row, COL["depth_mm"])
        h = _v(row, COL["height_mm"])
        dims_from_cols = None
        if any([w, d, h]):
            parts = [f"{x} mm" for label, x in [("W", w), ("D", d), ("H", h)] if x]
            dims_from_cols = " × ".join(parts) if parts else None

        product = ProductInput(
            sku_id=sku,
            product_name=name,
            room=_v(row, COL["room"]),
            product_type=_v(row, COL["product_type"]),
            imported=_parse_imported(_v(row, COL["imported"])),
            existing_small_description=_v(row, COL["small_description"]),
            existing_design_story=_v(row, COL["design_story"]),
            existing_meta_keywords=_v(row, COL["meta_keywords"]),
            existing_bullet_points=_v(row, COL["bullet_points"]),
            primary_color=_v(row, COL["primary_color"]),
            secondary_color=_v(row, COL["secondary_color"]),
            finish=_v(row, COL["finish"]),
            primary_material=_v(row, COL["primary_material"]),
            filling_material=_v(row, COL["filling_material"]),
            upholstery_type=_v(row, COL["upholstery_type"]),
            leg_material=_v(row, COL["leg_material"]),
            configuration=_v(row, COL["configuration"]),
            dimensions=_v(row, COL["dimensions"]) or dims_from_cols,
            net_weight=_v(row, COL["net_weight"]),
            width_mm=w,
            depth_mm=d,
            height_mm=h,
            max_load_capacity=_v(row, COL["max_load_capacity"]),
            mechanism=_v(row, COL["mechanism"]),
            function=_v(row, COL["function"]),
            warranty=_v(row, COL["warranty"]),
            style=_v(row, COL["style"]),
            care_instructions=_v(row, COL["care_instructions"]),
            w_icon_1=_v(row, COL["w_icon_1"]),
            w_icon_2=_v(row, COL["w_icon_2"]),
            w_icon_3=_v(row, COL["w_icon_3"]),
            w_icon_4=_v(row, COL["w_icon_4"]),
        )
        products.append(product)

    wb.close()
    return products


# ── MiracleAi gold example loader ─────────────────────────────────────────────

def load_miracle_ai_examples(file_path: str) -> List[Dict]:
    """
    Parse MiracleAi.xlsx into a list of gold example dicts.
    Columns: Category (0), SKU (1), Design Story (2), WYNI (3), WYLI Icons (4).
    Returns only rows that have at least a Design Story.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    examples = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        category  = str(row[0]).strip() if row[0] else ""
        sku       = str(row[1]).strip() if row[1] else ""
        story     = str(row[2]).strip() if row[2] else ""
        wyni      = str(row[3]).strip() if row[3] else ""
        wyli      = str(row[4]).strip() if row[4] else ""

        if not story:
            continue

        examples.append({
            "category":     category,
            "sku":          sku,
            "design_story": story,
            "what_you_need_to_know": wyni,
            "wyli_icon_text": wyli,
        })

    wb.close()
    return examples


def get_column_mapping_report(file_path: str) -> Dict:
    """Summary of what was detected in the file — shown in the UI after upload."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]

    header_idx = _find_header_row(ws)
    header_row = list(ws.iter_rows(min_row=header_idx + 1, max_row=header_idx + 1, values_only=True))[0]

    mapped = {}
    for field, col_idx in COL.items():
        if col_idx < len(header_row) and header_row[col_idx]:
            mapped[field] = str(header_row[col_idx]).strip()

    wb.close()
    return {
        "total_columns": len(header_row),
        "header_row": header_idx + 1,
        "mapped_fields": mapped,
        "coverage_pct": round(len(mapped) / len(COL) * 100),
    }
